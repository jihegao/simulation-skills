"""Aircraft mission reliability DES model driven by an external mission-config workbook.

The model is compatible with the earlier interface and also accepts a task
configuration Excel path. It models:
- scheduled mission waves derived from 任务剖面/基本任务
- stochastic in-flight failures from 故障建模
- maintenance crew queue contention (保障资源: 人员)
- parts stock consumption and shortage delay from 保障资源
- optional preventive maintenance windows
"""

from __future__ import annotations

import math
import random
import statistics
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

import openpyxl
import simpy


class AircraftMissionReliabilityModel:
    """Run mission reliability simulation with maintenance-resource coupling."""

    def __init__(
        self,
        *,
        seed: int = 2026,
        sample_interval: float = 0.25,
        horizon: float = 48.0,
        mission_interval: float = 2.0,
        mission_time: float = 2.0,
        mission_wave_size: int = 2,
        mtbf: float = 24.0,
        mttr: float = 2.5,
        part_fill_rate: float = 0.9,
        part_delay: float = 24.0,
        pm_interval: float = 25.0,
        pm_time: float = 4.0,
        maintenance_personnel: int = 10,
        mission_config_path: Optional[str] = None,
    ) -> None:
        self.seed = int(seed)
        self.rng = random.Random(self.seed)

        self.sample_interval = float(sample_interval)
        self.horizon = float(horizon)
        self.part_delay = float(part_delay)
        self.maintenance_personnel = max(1, int(maintenance_personnel))

        self.legacy_mission_interval = float(mission_interval)
        self.legacy_mission_time = float(mission_time)
        self.legacy_mission_wave_size = int(mission_wave_size)
        self.legacy_num_aircraft = 5

        self.mtbf = float(mtbf)
        self.mttr = float(mttr)
        self.pm_interval = float(pm_interval)
        self.pm_time = float(pm_time)
        self.part_fill_rate = float(part_fill_rate)

        if self.mtbf <= 0:
            raise ValueError("mtbf must be > 0")
        if self.sample_interval <= 0:
            raise ValueError("sample_interval must be > 0")
        if self.horizon <= 0:
            raise ValueError("horizon must be > 0")

        self.env = simpy.Environment()
        self.crew: Optional[simpy.Resource] = None
        self.maintenance_crew: Optional[simpy.Resource] = None

        # Runtime counters
        self.rows: List[Dict[str, Any]] = []
        self._task_rows: List[Dict[str, Any]] = []

        self.planned_missions = 0
        self.dispatched_missions = 0
        self.successful_missions = 0
        self.failed_missions = 0
        self.canceled_missions = 0

        self.pm_events = 0
        self.repair_events = 0
        self.part_shortage_events = 0
        self.part_wait_total = 0.0
        self.crew_wait_total = 0.0
        self.parts_consumed = 0

        self.crew_queue_max = 0
        self.parts_queue_max = 0

        # Config-driven entities
        self.mission_config_path = mission_config_path
        self.aircraft: List[Dict[str, Any]] = []
        self.mission_schedule: List[Dict[str, Any]] = []
        self.failure_catalog: Dict[str, List[Dict[str, Any]]] = {}
        self.spare_parts: Dict[str, int] = {}
        self.pm_setting = {
            "interval_h": self.pm_interval,
            "duration_h": self.pm_time,
        }

        if mission_config_path:
            self._load_config(Path(mission_config_path))
        else:
            self._build_legacy_config()

        self.crew = simpy.Resource(self.env, capacity=self.maintenance_personnel)
        self.maintenance_crew = self.crew

    def _to_float(self, value: Any, default: float = 0.0) -> float:
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).strip())
        except Exception:
            return default

    def _to_int(self, value: Any, default: int = 0) -> int:
        try:
            return int(float(self._to_float(value, float(default))))
        except Exception:
            return default

    def _as_str(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _load_sheet_rows(self, ws: Any) -> List[Dict[str, Any]]:
        rows = []
        header = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [self._as_str(v) for v in row]
                continue
            if not any(v is not None for v in row):
                continue
            rows.append({k: row[idx] for idx, k in enumerate(header)})
        return rows

    def _infer_hours(self, raw_mtbf: float) -> float:
        if raw_mtbf is None:
            return 0.0
        if raw_mtbf > 1000:
            # most values in workbook are "minutes" order of magnitude.
            return raw_mtbf / 60.0
        if raw_mtbf > 0 and raw_mtbf < 0.1:
            # already converted small values are treated as hours, keep as is.
            return raw_mtbf * 60.0
        return float(raw_mtbf)

    def _build_legacy_config(self) -> None:
        profiles = [
            {
                "mission_id": 1,
                "name": "legacy",
                "aircraft_type": "XX5",
                "aircraft_base": "pudong",
                "count": self.legacy_mission_wave_size,
                "min_launch": self.legacy_mission_wave_size,
                "prep_time": 0.0,
                "duration": self.legacy_mission_time,
                "countdown": self.horizon / self.legacy_mission_interval + 1,
                "interval": self.legacy_mission_interval,
            }
        ]
        self.mission_schedule = []
        for entry in profiles:
            for i in range(int(entry["countdown"])):
                t = i * entry["interval"]
                if t > self.horizon:
                    break
                mid = int(entry["mission_id"] * 100 + i * 10)
                self.mission_schedule.append(
                    {
                        "time": float(t),
                        "mission_id": mid,
                        "mission_name": entry["name"],
                        "aircraft_type": entry["aircraft_type"],
                        "aircraft_base": entry["aircraft_base"],
                        "count": entry["count"],
                        "min_launch": entry["min_launch"],
                        "duration": entry["duration"],
                        "prep_time": entry["prep_time"],
                    }
                )
        self.mission_schedule.sort(key=lambda x: (x["time"], x["mission_id"]))

        self.aircraft = []
        for idx in range(self.legacy_num_aircraft):
            self.aircraft.append(
                {
                    "aircraft_id": idx,
                    "aircraft_type": "XX5",
                    "aircraft_base": "pudong",
                    "status": "idle",
                    "busy_until": 0.0,
                    "next_pm_due": self.pm_interval,
                    "pm_pending": False,
                    "pm_interval": self.pm_interval,
                }
            )
        self.failure_catalog["XX5"] = [
            {
                "name": "generic_component",
                "quantity": 1,
                "mtbf_h": self.mtbf,
                "mttr_h": self.mttr,
            }
        ]
        self.spare_parts = {}

    def _load_config(self, path: Path) -> None:
        if not path.exists():
            raise FileNotFoundError(f"mission config not found: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)

        deployment = self._load_sheet_rows(wb["飞机部署"]) if "飞机部署" in wb.sheetnames else []
        failure_rows = self._load_sheet_rows(wb["故障建模"]) if "故障建模" in wb.sheetnames else []
        tasks = self._load_sheet_rows(wb["基本任务"]) if "基本任务" in wb.sheetnames else []
        profiles = self._load_sheet_rows(wb["任务剖面"]) if "任务剖面" in wb.sheetnames else []
        spare_rows = self._load_sheet_rows(wb["保障资源"]) if "保障资源" in wb.sheetnames else []
        pm_rows = self._load_sheet_rows(wb["预防性维修"]) if "预防性维修" in wb.sheetnames else []
        support_rows = self._load_sheet_rows(wb["保障点"]) if "保障点" in wb.sheetnames else []

        # Aircraft fleet.
        aid = 0
        self.aircraft = []
        for r in deployment:
            craft_type = self._as_str(r.get("机型"))
            base = self._as_str(r.get("机场"))
            count = self._to_int(r.get("数量"), 0)
            for _ in range(max(0, count)):
                self.aircraft.append(
                    {
                        "aircraft_id": aid,
                        "aircraft_type": craft_type,
                        "aircraft_base": base,
                        "status": "idle",
                        "busy_until": 0.0,
                        "next_pm_due": self.pm_interval,
                        "pm_pending": False,
                        "pm_interval": self.pm_interval,
                    }
                )
                aid += 1

        if not self.aircraft:
            # Fallback to legacy setup.
            self._build_legacy_config()
            return

        # Failure model by type and component.
        self.failure_catalog = {}
        for r in failure_rows:
            craft_type = self._as_str(r.get("装备名称"))
            if not craft_type:
                continue
            name = self._as_str(r.get("组件名称"))
            qty = self._to_int(r.get("数量"), 1)
            raw_mtbf = self._to_float(r.get("MTBF"), self.mtbf)
            raw_mttr = self._to_float(r.get("MTTR/min"), self.mttr * 60.0)
            mtbf_h = self._infer_hours(raw_mtbf)
            mttr_h = max(0.05, raw_mttr / 60.0)

            self.failure_catalog.setdefault(craft_type, []).append(
                {
                    "name": name,
                    "quantity": max(1, qty),
                    "mtbf_h": max(0.01, mtbf_h),
                    "mttr_h": mttr_h,
                }
            )

        # If a configured type has no failure rows, inherit legacy single-line profile.
        if not self.failure_catalog:
            self.failure_catalog[craft_type] = [
                {"name": "generic_component", "quantity": 1, "mtbf_h": self.mtbf, "mttr_h": self.mttr}
                for craft_type in {a["aircraft_type"] for a in self.aircraft}
            ]

        # Mission schedule from task profile.
        profile_map = {}
        for p in profiles:
            name = self._as_str(p.get("任务剖面"))
            if not name:
                continue
            profile_map[name] = {
                "start": self._to_float(p.get("开始时间"), 0.0),
                "interval": self._to_float(p.get("重复周期"), 0.0),
                "repeat": self._to_int(p.get("重复数量"), 0),
                "unit": self._as_str(p.get("时间单位")).upper() if p.get("时间单位") is not None else "HOUR",
            }

        self.mission_schedule = []
        for t in tasks:
            name = self._as_str(t.get("任务名称"))
            base_profile = profile_map.get(name, {"start": 0.0, "interval": 0.0, "repeat": 0, "unit": "HOUR"})
            start = self._to_float(base_profile.get("start"), 0.0)
            interval = self._to_float(base_profile.get("interval"), 0.0)
            repeat = self._to_int(base_profile.get("repeat"), 0)
            unit = self._as_str(base_profile.get("unit")) or "HOUR"
            interval_h = interval
            if "DAY" in unit.upper():
                interval_h *= 24.0

            task_type = self._as_str(t.get("装备类型"))
            base = self._as_str(t.get("出动机场"))
            mission_count = max(0, self._to_int(t.get("装备数量"), 0))
            min_launch = max(0, self._to_int(t.get("最低出动数量"), mission_count))
            prep = self._to_float(t.get("提前准备时间"), 0.0)
            duration = self._to_float(t.get("任务时长_小时"), self.legacy_mission_time)
            mid = max(1, self._to_int(t.get("ID"), 1))

            for k in range(max(0, repeat)):
                t0 = start + k * interval_h
                if t0 > self.horizon:
                    continue
                self.mission_schedule.append(
                    {
                        "time": t0,
                        "mission_id": int(mid * 1000 + k),
                        "mission_name": name or f"任务{mid}",
                        "aircraft_type": task_type,
                        "aircraft_base": base,
                        "count": mission_count,
                        "min_launch": min_launch,
                        "duration": duration,
                        "prep_time": prep,
                    }
                )

        if not self.mission_schedule:
            # fallback to legacy profile if config had empty task list.
            self._build_legacy_config()
            return

        self.mission_schedule.sort(key=lambda x: (x["time"], x["mission_id"]))

        # Spare parts at support points and total available repair crew.
        crew_total = 0
        for row in support_rows:
            if self._as_str(row.get("类型")) == "人员":
                crew_total += self._to_int(row.get("数量"), 0)
        if crew_total > 0:
            self.maintenance_personnel = crew_total

        self.spare_parts = {}
        for row in spare_rows:
            if self._as_str(row.get("类型")) != "备件":
                continue
            part = self._as_str(row.get("资源"))
            qty = self._to_int(row.get("数量"), 0)
            if not part:
                continue
            self.spare_parts[part] = self.spare_parts.get(part, 0) + max(0, qty)

        # Optional PM information from config.
        if pm_rows:
            # use the first row if it matches known aircraft pattern; fallback to defaults.
            first = pm_rows[0]
            pm_hour = self._to_float(first.get("维修小时"), self.pm_time)
            interval_1 = self._to_float(first.get("预防性维修周期_H"), self.pm_interval)
            if pm_hour > 0:
                self.pm_setting["duration_h"] = pm_hour
            if interval_1 > 0:
                self.pm_setting["interval_h"] = interval_1

            for state in self.aircraft:
                state["pm_interval"] = self.pm_setting["interval_h"]
                state["next_pm_due"] = self.pm_setting["interval_h"]
        
    def _record_sample(self) -> None:
        now = self.env.now
        status_counts = {"idle": 0, "mission": 0, "pm": 0, "repair": 0}
        for state in self.aircraft:
            status_counts[str(state["status"]) ] = status_counts.get(str(state["status"]), 0) + 1

        crew_waiters = 0
        if self.maintenance_crew is not None:
            crew_waiters = len(self.maintenance_crew.queue)
            self.crew_queue_max = max(self.crew_queue_max, crew_waiters)

        part_waiters = 0
        for count in self.spare_parts.values():
            if count <= 0:
                part_waiters += 1

        self.rows.append(
            {
                "time": round(now, 3),
                "kind": 2,
                "sample_idle": float(status_counts.get("idle", 0)),
                "sample_mission": float(status_counts.get("mission", 0)),
                "sample_pm": float(status_counts.get("pm", 0)),
                "sample_repair": float(status_counts.get("repair", 0)),
                "sample_crew_wait": float(crew_waiters),
                "sample_parts_shortage": float(part_waiters),
                "sample_available": float(status_counts.get("idle", 0)),
                "sample_planned": self.planned_missions,
                "sample_dispatched": self.dispatched_missions,
                "sample_successful": self.successful_missions,
            }
        )

    def _sampler(self) -> simpy.events.Process:
        while True:
            self._record_sample()
            yield self.env.timeout(self.sample_interval)

    def _draw_failure(self, state: Dict[str, Any]) -> tuple[float, Dict[str, Any]]:
        craft_type = state["aircraft_type"]
        components = self.failure_catalog.get(craft_type, [])
        if not components:
            # legacy fallback.
            return float("inf"), {"name": "generic", "mttr_h": self.mttr}

        best_dt = float("inf")
        best_component = components[0]
        for item in components:
            mtbf_h = max(1e-6, float(item.get("mtbf_h", self.mtbf)))
            qty = max(1, int(item.get("quantity", 1)))
            rate = qty / mtbf_h
            dt = self.rng.expovariate(rate)
            if dt < best_dt:
                best_dt = dt
                best_component = item
        return best_dt, best_component

    def _start_preventive_maintenance(self, state: Dict[str, Any], now: float) -> bool:
        pm_interval = float(state.get("pm_interval", 0.0))
        if pm_interval <= 0:
            return False
        if state["status"] != "idle" or float(state["next_pm_due"]) > now:
            return False

        self.pm_events += 1
        due = float(state["next_pm_due"])
        state["next_pm_due"] = due + pm_interval

        def _pm_task() -> simpy.events.Process:
            state["status"] = "pm"
            self.rows.append(
                {
                    "time": round(self.env.now, 3),
                    "kind": 6,
                    "aircraft_id": int(state["aircraft_id"]),
                    "event": "pm_start",
                    "duration": round(self.pm_setting["duration_h"], 3),
                    "due_time": round(due, 3),
                }
            )
            with self.maintenance_crew.request() as req:
                t0 = self.env.now
                yield req
                queue_wait = self.env.now - t0
                if queue_wait > 0:
                    self.crew_wait_total += queue_wait
                yield self.env.timeout(self.pm_setting["duration_h"])
            state["status"] = "idle"
            self.rows.append(
                {
                    "time": round(self.env.now, 3),
                    "kind": 6,
                    "aircraft_id": int(state["aircraft_id"]),
                    "event": "pm_end",
                    "duration": round(self.pm_setting["duration_h"], 3),
                }
            )

        self.env.process(_pm_task())
        return True

    def _ensure_pm_after_busy(self, state: Dict[str, Any], now: float) -> None:
        if float(state.get("next_pm_due", float("inf"))) <= now and state["status"] == "idle":
            self._start_preventive_maintenance(state, now)
        elif float(state.get("next_pm_due", float("inf"))) <= now:
            state["pm_pending"] = True

    def _start_repair(self, state: Dict[str, Any], now: float, component_name: str, repair_mean_h: float) -> None:
        if state["status"] == "repair":
            return

        self.repair_events += 1
        state["status"] = "repair"

        self.rows.append(
            {
                "time": round(now, 3),
                "kind": 6,
                "aircraft_id": int(state["aircraft_id"]),
                "event": "repair_start",
                "parts_short": 0,
                "parts_delay": 0.0,
                "component": component_name,
            }
        )

        def _repair_task() -> simpy.events.Process:
            # maintenance crew queue + repair.
            with self.maintenance_crew.request() as req:
                t0 = self.env.now
                yield req
                crew_wait = self.env.now - t0
                if crew_wait > 0:
                    self.crew_wait_total += crew_wait

                delay = 0.0
                if component_name:
                    stock = self.spare_parts.get(component_name, 0)
                    if stock > 0:
                        self.spare_parts[component_name] = stock - 1
                        self.parts_consumed += 1
                    else:
                        self.part_shortage_events += 1
                        self.parts_queue_max = max(self.parts_queue_max, sum(1 for v in self.spare_parts.values() if v <= 0))
                        delay = self.part_delay
                        self.part_wait_total += delay
                        self.rows[-1]["parts_short"] = 1
                        self.rows[-1]["parts_delay"] = round(delay, 3)
                        yield self.env.timeout(delay)
                        # assume replenishment arrives at this event.
                        self.spare_parts[component_name] = 0

                repair_time = max(0.05, self.rng.expovariate(1.0 / max(0.05, repair_mean_h)))
                yield self.env.timeout(repair_time)
                state["status"] = "idle"
                state["busy_until"] = self.env.now
                self._ensure_pm_after_busy(state, self.env.now)
                self.rows.append(
                    {
                        "time": round(self.env.now, 3),
                        "kind": 6,
                        "aircraft_id": int(state["aircraft_id"]),
                        "event": "repair_end",
                        "parts_delay": round(delay, 3),
                    }
                )

        self.env.process(_repair_task())

    def _perform_mission(
        self,
        aircraft_id: int,
        start_time: float,
        mission_id: int,
        mission_meta: Dict[str, Any],
    ) -> simpy.events.Process:
        state = self.aircraft[aircraft_id]
        duration = float(mission_meta["duration"])

        if state["status"] != "idle" or state["busy_until"] > start_time:
            return self.env.timeout(0)

        # If PM is due at/ before start, PM takes priority.
        if self._start_preventive_maintenance(state, start_time):
            # PM is scheduled and this sortie should be retried as canceled.
            self._record_mission_event(
                mission_id,
                state,
                start_time,
                0,
                0,
                "failed",
                "pm_blocked",
            )
            return self.env.timeout(0)

        if state["status"] != "idle" or state["busy_until"] > start_time:
            self._record_mission_event(
                mission_id,
                state,
                start_time,
                0,
                0,
                "canceled",
                "unavailable",
            )
            return self.env.timeout(0)

        state["status"] = "mission"
        self.dispatched_missions += 1

        fail_time, comp = self._draw_failure(state)
        fail_at = start_time + fail_time
        mission_end = start_time + duration

        if fail_at < mission_end:
            run_time = max(0.0, min(duration, fail_at - start_time))
            yield self.env.timeout(run_time)
            self.failed_missions += 1
            self._start_repair(state, self.env.now, str(comp.get("name", "unknown")), float(comp.get("mttr_h", self.mttr)))
            self._record_mission_event(
                mission_id,
                state,
                start_time,
                1,
                0,
                "failed",
                "mission_failed",
                task_time=round(self.env.now - start_time, 3),
                service_time=round(run_time, 3),
                parts_wait=0.0,
            )
            return

        # mission complete.
        yield self.env.timeout(duration)
        self.successful_missions += 1
        state["status"] = "idle"
        state["busy_until"] = self.env.now
        self._ensure_pm_after_busy(state, self.env.now)
        self._record_mission_event(
            mission_id,
            state,
            start_time,
            1,
            1,
            "success",
            "mission_success",
            task_time=round(duration, 3),
            service_time=round(duration, 3),
            parts_wait=0.0,
        )

    def _record_mission_event(
        self,
        mission_id: int,
        state: Dict[str, Any],
        plan_start: float,
        dispatched: int,
        success: int,
        result: str,
        event: str,
        task_time: float = 0.0,
        service_time: float = 0.0,
        parts_wait: float = 0.0,
    ) -> None:
        row = {
            "time": round(self.env.now, 3),
            "kind": 1,
            "aircraft_id": int(state["aircraft_id"]) if state else -1,
            "mission_id": int(mission_id),
            "plan_start": round(plan_start, 3),
            "queue_wait": 0.0,
            "service_time": round(service_time, 3),
            "parts_wait": round(parts_wait, 3),
            "task_time": round(task_time, 3),
            "mission_dispatched": int(dispatched),
            "mission_success": int(success),
            "mission_result": result,
            "event": event,
        }
        self.rows.append(row)
        self._task_rows.append(row)

    def _canceled_mission_event(self, mission_meta: Dict[str, Any], mission_id: int, plan_start: float) -> None:
        self.canceled_missions += 1
        self._task_rows.append(
            {
                "time": round(plan_start, 3),
                "kind": 5,
                "mission_id": int(mission_id),
                "aircraft_id": -1,
                "plan_start": round(plan_start, 3),
                "queue_wait": 0.0,
                "service_time": 0.0,
                "parts_wait": 0.0,
                "task_time": 0.0,
                "mission_dispatched": 0,
                "mission_success": 0,
                "mission_result": "canceled",
                "event": "mission_canceled",
                "mission_name": mission_meta.get("mission_name"),
            }
        )

    def _should_dispatch(self, now: float, base: str, craft_type: str) -> List[int]:
        candidates = [
            state
            for state in self.aircraft
            if state["status"] == "idle"
            and float(state["busy_until"]) <= now
            and state["aircraft_base"] == base
            and state["aircraft_type"] == craft_type
        ]
        candidates.sort(key=lambda item: int(item["aircraft_id"]))
        return [int(item["aircraft_id"]) for item in candidates]

    def _run_schedule(self) -> simpy.events.Process:
        def _runner() -> simpy.events.Process:
            for mission in self.mission_schedule:
                plan_t = float(mission["time"]) + float(mission.get("prep_time", 0.0))
                if plan_t > self.horizon:
                    break
                wait = plan_t - self.env.now
                if wait > 0:
                    yield self.env.timeout(wait)

                count = int(mission["count"])
                min_launch = int(mission["min_launch"])
                mission_id = int(mission["mission_id"])
                aircrafts = self._should_dispatch(self.env.now, mission["aircraft_base"], mission["aircraft_type"])
                self.planned_missions += count

                if not aircrafts:
                    for k in range(count):
                        self.canceled_missions += 1
                        self._canceled_mission_event(mission, mission_id * 10 + k, plan_t)
                    continue

                # If available aircraft below minimum launch requirement, we cancel all scheduled sorties.
                if len(aircrafts) < min_launch:
                    for slot in range(len(aircrafts)):
                        self.canceled_missions += 1
                        self._canceled_mission_event(mission, mission_id * 10 + slot, plan_t)
                    for k in range(count - len(aircrafts)):
                        self.canceled_missions += 1
                        self._canceled_mission_event(mission, mission_id * 10 + 100 + k, plan_t)
                    continue

                # dispatch all planned sorties up to available.
                for slot in range(min(count, len(aircrafts))):
                    aid = aircrafts[slot]
                    self.env.process(self._perform_mission(aid, self.env.now, mission_id * 10 + slot, mission))

                if len(aircrafts) < count:
                    for k in range(count - len(aircrafts)):
                        self.canceled_missions += 1
                        self._canceled_mission_event(mission, mission_id * 10 + 1000 + k, plan_t)

            # Keep sampling until run end.
        return _runner()

    def _summary_rows(self, until: float) -> List[Dict[str, Any]]:
        if self._task_rows:
            waits = [float(r.get("queue_wait", 0.0)) for r in self._task_rows]
            task_times = [float(r.get("task_time", 0.0)) for r in self._task_rows if float(r.get("mission_dispatched", 0.0)) > 0]
            p95_wait = statistics.quantiles(waits, n=20)[18] if len(waits) >= 2 else (waits[0] if waits else 0.0)
            p95_task = statistics.quantiles(task_times, n=20)[18] if len(task_times) >= 2 else (task_times[0] if task_times else 0.0)
            avg_wait = statistics.mean(waits) if waits else 0.0
            avg_task = statistics.mean(task_times) if task_times else 0.0
            p95_dispatched_queue = (
                statistics.quantiles([r.get("sample_crew_wait", 0.0) for r in self.rows if int(r.get("kind", 0)) == 2], n=20)[18]
                if sum(1 for r in self.rows if int(r.get("kind", 0)) == 2) >= 2
                else 0.0
            )
        else:
            p95_wait = 0.0
            p95_task = 0.0
            avg_wait = 0.0
            avg_task = 0.0
            p95_dispatched_queue = 0.0

        if self.planned_missions > 0:
            mission_reliability = self.successful_missions / self.planned_missions
            sortie_rate = self.dispatched_missions / self.planned_missions
            completion_rate = self.successful_missions / max(1, self.dispatched_missions)
        else:
            mission_reliability = 0.0
            sortie_rate = 0.0
            completion_rate = 0.0

        sample_rows = [r for r in self.rows if int(r.get("kind", 0)) == 2]
        if sample_rows:
            avg_idle = statistics.mean(float(r["sample_idle"]) for r in sample_rows)
            avg_mission = statistics.mean(float(r["sample_mission"]) for r in sample_rows)
            avg_pm = statistics.mean(float(r["sample_pm"]) for r in sample_rows)
            avg_repair = statistics.mean(float(r["sample_repair"]) for r in sample_rows)
            avg_crew_wait = statistics.mean(float(r.get("sample_crew_wait", 0.0)) for r in sample_rows)
        else:
            avg_idle = avg_mission = avg_pm = avg_repair = avg_crew_wait = 0.0

        mission_gap = 0.0
        if len(self._task_rows) >= 2:
            start_times = [float(r["plan_start"]) for r in self._task_rows if int(r.get("mission_dispatched", 0))]
            if len(start_times) >= 2:
                diffs = [j - i for i, j in zip(start_times[:-1], start_times[1:]) if j >= i]
                if diffs:
                    mission_gap = statistics.mean(diffs)

        summary = {
            "time": round(until, 3),
            "kind": 3,
            "summary_planned_missions": int(self.planned_missions),
            "summary_dispatched_missions": int(self.dispatched_missions),
            "summary_successful_missions": int(self.successful_missions),
            "summary_failed_missions": int(self.failed_missions),
            "summary_canceled_missions": int(self.canceled_missions),
            "summary_mission_reliability": round(mission_reliability, 6),
            "summary_sortie_rate": round(sortie_rate, 6),
            "summary_success_when_dispatched": round(completion_rate, 6),
            "summary_pm_events": int(self.pm_events),
            "summary_repair_events": int(self.repair_events),
            "summary_part_shortage_events": int(self.part_shortage_events),
            "summary_part_wait_total": round(self.part_wait_total, 3),
            "summary_seed": self.seed,
            "summary_mean_queue_wait": round(avg_wait, 3),
            "summary_p95_queue_wait": round(p95_wait, 3),
            "summary_task_time_p95": round(p95_task, 3),
            "summary_mean_task_time": round(avg_task, 3),
            "summary_mean_sample_idle": round(avg_idle, 3),
            "summary_mean_sample_mission": round(avg_mission, 3),
            "summary_mean_sample_pm": round(avg_pm, 3),
            "summary_mean_sample_repair": round(avg_repair, 3),
            "summary_mean_crew_wait": round(avg_crew_wait, 3),
            "summary_crew_wait_p95": round(p95_dispatched_queue, 3),
            "summary_parts_queue_max": int(self.parts_queue_max),
            "summary_crew_queue_max": int(self.crew_queue_max),
            "summary_crew_wait_total": round(self.crew_wait_total, 3),
            "summary_mean_gap": round(mission_gap, 3),
            "summary_parts_level_end": int(sum(int(v) for v in self.spare_parts.values())),
            "summary_parts_consumed": int(self.parts_consumed),
        }

        return [summary]

    def run(self, until: float) -> List[Dict[str, Any]]:
        if until <= 0:
            until = self.horizon

        self.rows = []
        self._task_rows = []
        for state in self.aircraft:
            state["status"] = "idle"
            state["busy_until"] = 0.0
            state["next_pm_due"] = float(state.get("pm_interval", self.pm_interval))
            state["pm_pending"] = False

        self.env.process(self._sampler())
        self.env.process(self._run_schedule())
        self.env.run(until=until)

        self.rows.extend(self._summary_rows(until))
        return self.rows
